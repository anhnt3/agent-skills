#!/usr/bin/env python3
"""CSV → XLSX cho testcase QA (2 sheet: Testcases + Ma trận truy vết).
Agnostic: chỉ format CSV theo hợp đồng 16 cột. Tự dựng venv + openpyxl lần đầu.
Chỉ cần python3."""
from __future__ import annotations
import csv
import json
import os
import subprocess
import sys
from pathlib import Path

EXPECTED_HEADER = [
    "ID", "Tiêu đề", "Nhóm", "Ưu tiên", "Loại", "Tiền điều kiện",
    "Dữ liệu test", "Các bước thực hiện", "Kết quả mong đợi", "Truy vết",
    "Test tự động", "Kết quả tự động",
    "Kết quả thực tế", "Trạng thái", "Bug ID", "Ghi chú",
]
COL_ID = 0
COL_TRACE = 9
COL_AUTO_TEST = 10
COL_AUTO_STATUS = 11
STATUS_CHOICES = ["Pass", "Fail", "Blocked", "N/A", "Chưa chạy"]
MANUAL_ONLY_SENTINEL = "manual-only"


def read_cases(csv_path):
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    if not rows:
        raise ValueError("CSV rỗng")
    return rows[0], rows[1:]


def read_cases_json(json_path):
    """Đọc testcase từ JSON (mảng object, mỗi object đủ 16 key = EXPECTED_HEADER).
    Ưu điểm so với CSV: không cần tự quote/escape newline, dấu phẩy trong tiếng Việt.
    Các cột thực thi (13-16: Kết quả thực tế, Trạng thái, Bug ID, Ghi chú) vẫn phải
    có mặt như key nhưng giá trị nên để chuỗi rỗng "" (tester điền sau khi chạy tay)."""
    with open(json_path, encoding="utf-8-sig") as f:
        cases = json.load(f)
    if not isinstance(cases, list):
        raise ValueError("JSON phải là một mảng (list) các object testcase")
    expected_keys = set(EXPECTED_HEADER)
    rows = []
    for i, case in enumerate(cases):
        if not isinstance(case, dict):
            raise ValueError(f"Case {i}: phải là object, nhận {type(case).__name__}")
        keys = set(case.keys())
        missing = expected_keys - keys
        extra = keys - expected_keys
        if missing or extra:
            parts = []
            if missing:
                parts.append("thiếu key " + ", ".join(f"'{k}'" for k in sorted(missing)))
            if extra:
                parts.append("thừa key " + ", ".join(f"'{k}'" for k in sorted(extra)))
            raise ValueError(f"Case {i}: " + "; ".join(parts))
        rows.append([case[h] for h in EXPECTED_HEADER])
    return list(EXPECTED_HEADER), rows


def validate_header(header):
    if list(header) != EXPECTED_HEADER:
        raise ValueError(
            f"Header sai. Cần {len(EXPECTED_HEADER)} cột: {EXPECTED_HEADER}; nhận: {header}")


def validate_rows(header, rows):
    n = len(EXPECTED_HEADER)
    bad = []
    for i, r in enumerate(rows, start=2):
        if len(r) != n:
            id_ = f" (ID={r[0]!r})" if r else ""
            bad.append(f"Dòng {i}: {len(r)} field, cần {n}{id_}")
    if bad:
        raise ValueError("Có dòng sai số cột:\n" + "\n".join(bad))


def parse_traceability(cell):
    if not cell or not cell.strip():
        return []
    parts = [p.strip() for chunk in cell.split(",") for p in chunk.split(";")]
    return [p for p in parts if p]


def aggregate_status(statuses):
    if not statuses:
        return "chưa chạy"

    vals = [s.strip() for s in statuses]
    non_empty = [v for v in vals if v]

    # If no non-empty values, return "chưa chạy"
    if not non_empty:
        return "chưa chạy"

    # If there are empty values mixed with non-empty, return "chưa chạy"
    if len(non_empty) < len(vals):
        return "chưa chạy"

    # Among non-empty values, separate real (non-"N/A") results.
    real = [v for v in non_empty if v.lower() != "n/a"]

    # If any real result is a fail, return "Fail"
    if any(v.lower().startswith("fail") for v in real):
        return "Fail"

    # If there are real results and all pass, return "Pass"
    if real and all(v.lower().startswith("pass") for v in real):
        return "Pass"

    # All non-empty values were "N/A"
    if not real:
        return "N/A"

    return "chưa chạy"


def build_matrix(rows):
    order, bucket = [], {}
    for r in rows:
        for fr in parse_traceability(r[COL_TRACE]):
            if fr not in bucket:
                bucket[fr] = {
                    "fr": fr, "tcs": [], "auto_tests": [], "statuses": [],
                    "has_manual_sentinel": False,
                }
                order.append(fr)
            b = bucket[fr]
            if r[COL_ID].strip():
                b["tcs"].append(r[COL_ID].strip())
            auto_test_val = r[COL_AUTO_TEST].strip()
            if auto_test_val:
                if auto_test_val.lower() == MANUAL_ONLY_SENTINEL:
                    b["has_manual_sentinel"] = True
                else:
                    b["auto_tests"].append(auto_test_val)
            b["statuses"].append(r[COL_AUTO_STATUS])
    out = []
    for fr in order:
        b = bucket[fr]
        if b["auto_tests"]:
            cover = "Có"
        elif b["has_manual_sentinel"]:
            cover = "MANUAL"
        else:
            cover = "GAP"
        out.append({
            "fr": fr,
            "tcs": b["tcs"],
            "auto_tests": b["auto_tests"],
            "auto_status": aggregate_status(b["statuses"]),
            "cover": cover,
        })
    return out


def _safe_sheet(name):
    for ch in ':\\/?*[]':
        name = name.replace(ch, ' ')
    return name[:31] or "Testcases"


def _read_existing_exec(out_path, sheet_name="Testcases"):
    """Đọc cột 13-16 (thực thi của tester) từ file xlsx đã có, keyed theo ID (cột 1).
    Trả về {} nếu file chưa tồn tại hoặc sheet không tồn tại. Import openpyxl lazily."""
    if not Path(out_path).exists():
        return {}
    from openpyxl import load_workbook

    try:
        wb = load_workbook(out_path)
    except Exception:
        return {}
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    result = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=16, values_only=True):
        if not row:
            continue
        id_ = row[0]
        if id_ is None or not str(id_).strip():
            continue
        id_ = str(id_).strip()
        vals = []
        for idx in (12, 13, 14, 15):
            v = row[idx] if idx < len(row) else None
            vals.append("" if v is None else v)
        result[id_] = vals
    return result


def write_xlsx(header, rows, matrix, out_path, sheet_name="Testcases"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    safe_sheet_name = _safe_sheet(sheet_name)
    preserved = _read_existing_exec(out_path, safe_sheet_name)

    # ID là khóa merge cột 13-16 (dữ liệu tester). ID cũ không còn trong input mới
    # = dữ liệu tester của case đó bị bỏ. Cảnh báo thay vì mất im lặng.
    new_ids = {str(r[0] or "").strip() for r in rows}
    lost = sorted(
        i for i, vals in preserved.items()
        if i not in new_ids and any(str(v).strip() for v in vals)
    )
    if lost:
        print(
            f"WARNING: {len(lost)} ID có dữ liệu tester trong xlsx cũ nhưng không có "
            f"trong input mới, các ô thực thi sẽ mất: {', '.join(lost)}",
            file=sys.stderr,
        )

    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(bold=True, color="FFFFFF")
    auto_fill = PatternFill("solid", fgColor="E2EFDA")   # cột máy (auto)
    exec_fill = PatternFill("solid", fgColor="FFF2CC")   # cột người (thực thi)
    wrap = Alignment(wrap_text=True, vertical="top")
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    ws = wb.active
    ws.title = safe_sheet_name
    ws.append(list(header))
    for c in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font = header_fill, header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = border
    for r in rows:
        r = list(r)
        id_ = r[COL_ID].strip() if r[COL_ID] else ""
        pres = preserved.get(id_)
        if pres is not None:
            for offset, col_idx in enumerate((12, 13, 14, 15)):
                incoming = r[col_idx] if col_idx < len(r) else ""
                if not (incoming and str(incoming).strip()):
                    r[col_idx] = pres[offset]
        ws.append(r)
    widths = [12, 34, 16, 8, 14, 26, 22, 34, 34, 14, 30, 16, 24, 14, 12, 22]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    last = ws.max_row
    for row in ws.iter_rows(min_row=2, max_row=last, min_col=1, max_col=len(header)):
        for cell in row:
            cell.alignment = wrap
            cell.border = border
    for row in range(2, last + 1):
        for col in (COL_AUTO_TEST + 1, COL_AUTO_STATUS + 1):
            ws.cell(row=row, column=col).fill = auto_fill
        for col in range(13, 17):
            ws.cell(row=row, column=col).fill = exec_fill
    dv = DataValidation(type="list", formula1='"%s"' % ",".join(STATUS_CHOICES), allow_blank=True)
    ws.add_data_validation(dv)
    if last >= 2:
        dv.add(f"N2:N{last}")
    ws.freeze_panes = "A2"
    lr = last + 2
    ws.cell(row=lr, column=1, value="Chú thích:").font = Font(bold=True)
    ws.cell(row=lr, column=2, value="Cột xanh = kết quả tự động (chỉ đọc)").fill = auto_fill
    ws.cell(row=lr, column=4, value="Cột vàng = tester điền khi chạy tay").fill = exec_fill

    ms = wb.create_sheet("Ma trận truy vết")
    mheader = ["FR/AC", "Manual TC", "Test tự động", "Auto status", "Phủ"]
    ms.append(mheader)
    for c in range(1, len(mheader) + 1):
        cell = ms.cell(row=1, column=c)
        cell.fill, cell.font, cell.border = header_fill, header_font, border
    for mm in matrix:
        ms.append([
            mm["fr"], ", ".join(mm["tcs"]),
            "\n".join(mm["auto_tests"]) if mm["auto_tests"] else "—",
            mm["auto_status"], mm["cover"],
        ])
    for c, w in enumerate([16, 24, 44, 14, 8], start=1):
        ms.column_dimensions[get_column_letter(c)].width = w
    for row in ms.iter_rows(min_row=2, max_row=ms.max_row, min_col=1, max_col=len(mheader)):
        for cell in row:
            cell.alignment = wrap
            cell.border = border
    ms.freeze_panes = "A2"

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main(argv=None):
    import argparse
    argv = list(sys.argv[1:] if argv is None else argv)
    p = argparse.ArgumentParser(description="CSV/JSON → XLSX testcase (2 sheet)")
    p.add_argument("csv_path", help="File CSV (16 cột) hoặc JSON (mảng object 16 key)")
    p.add_argument("xlsx_path")
    p.add_argument("--sheet", default="Testcases")
    a = p.parse_args(argv)
    if a.csv_path.lower().endswith(".json"):
        header, rows = read_cases_json(a.csv_path)
    else:
        header, rows = read_cases(a.csv_path)
    validate_header(header)
    validate_rows(header, rows)
    matrix = build_matrix(rows)
    write_xlsx(header, rows, matrix, a.xlsx_path, a.sheet)
    print(f"OK: {len(rows)} case, {len(matrix)} FR/AC → {a.xlsx_path}")


def _has_openpyxl():
    try:
        import openpyxl  # noqa: F401
        return True
    except ImportError:
        return False


def _bootstrap_and_reexec():
    here = Path(__file__).resolve().parent
    venv = here / ".venv"
    py = venv / ("Scripts" if os.name == "nt" else "bin") / "python"
    if not py.exists():
        print("Lần đầu: đang tạo venv + cài openpyxl...", file=sys.stderr)
        subprocess.check_call([sys.executable, "-m", "venv", str(venv)])
        subprocess.check_call([str(py), "-m", "pip", "install", "-q", "openpyxl"])
    os.execv(str(py), [str(py), str(Path(__file__).resolve()), *sys.argv[1:]])


if __name__ == "__main__":
    if not _has_openpyxl() and os.environ.get("QA_XLSX_NO_BOOTSTRAP") != "1":
        _bootstrap_and_reexec()
    main()

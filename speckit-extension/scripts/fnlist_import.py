#!/usr/bin/env python3
"""Function list (.xlsx/.csv) → .specify/docs/functions.json.

Kỷ luật: SCRIPT CHÉP NGUYÊN VĂN, LLM chỉ quyết ánh xạ cột và kiểu phân cấp.
Đây là văn bản hợp đồng nghiệm thu — không tóm tắt, không chuẩn hoá, không
"làm đẹp" nội dung ô. Script là thứ DUY NHẤT ghi functions.json.

Ba subcommand:
  inspect  — in cấu trúc thật của file + ứng viên kiểu phân cấp. Không đoán gì.
  write    — nhận ánh xạ cột dạng JSON, dựng cây, cấp ID, ghi functions.json.
  update   — đổi `status` của một/nhiều FN-ID (code-intel gọi khi ghi ngược).

Logic dựng cây/cấp ID nằm ở fnlist_tree.py. Tự dựng venv + openpyxl lần đầu
(chỉ khi đọc .xlsx). Chỉ cần python3.
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import subprocess
import sys
import tempfile
from pathlib import Path

import fnlist_tree as ft

def _force_utf8_console() -> None:
    """Windows mở subprocess với stdout/stderr mã cp1252 mặc định, không phải
    UTF-8 — in tiếng Việt có dấu ra đó là crash UnicodeEncodeError. Ép lại UTF-8
    khi có thể; capsys của pytest thay stdout bằng đối tượng không có
    `reconfigure` nên phải hasattr trước."""
    for name in ("stdout", "stderr"):
        stream = getattr(sys, name, None)
        enc = getattr(stream, "encoding", None)
        if enc and enc.lower() != "utf-8" and hasattr(stream, "reconfigure"):
            stream.reconfigure(encoding="utf-8", errors="replace")


_force_utf8_console()

XLSX_SUFFIXES = {".xlsx", ".xlsm"}


def cell_str(v) -> str:
    """Một ô → chuỗi. Số nguyên dạng float (openpyxl trả 3.0) về "3" cho khỏi
    lệch với bản gõ tay trong Word/Excel."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _read_csv(path: Path) -> list[list[str]]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        return [[cell_str(c) for c in row] for row in csv.reader(f)]


def _read_xlsx(path: Path, sheet: str | None) -> dict[str, list[list[str]]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    names = [sheet] if sheet else wb.sheetnames
    out: dict[str, list[list[str]]] = {}
    for name in names:
        if name not in wb.sheetnames:
            raise SystemExit(f"Không có sheet '{name}'. Sheet hiện có: {wb.sheetnames}")
        ws = wb[name]
        rows = [[cell_str(c) for c in row] for row in ws.iter_rows(values_only=True)]
        while rows and not any(rows[-1]):
            rows.pop()
        out[name] = rows
    return out


def read_grid(path, sheet: str | None = None) -> dict[str, list[list[str]]]:
    path = Path(path)
    if not path.exists():
        raise SystemExit(f"Không thấy file: {path}")
    if path.suffix.lower() in XLSX_SUFFIXES:
        return _read_xlsx(path, sheet)
    return {path.stem: _read_csv(path)}


def cmd_inspect(a) -> None:
    grids = read_grid(a.path, a.sheet)
    out = {"file": str(a.path), "sheets": []}
    for name, rows in grids.items():
        out["sheets"].append({
            "name": name,
            "rows": len(rows),
            "cols": max((len(r) for r in rows), default=0),
            "head": [r[: a.max_cols] for r in rows[: a.max_rows]],
            # Ứng viên kiểu phân cấp — chỉ là phỏng đoán kèm bằng chứng, LLM
            # BẮT BUỘC hỏi người dùng xác nhận chứ không tự chọn ứng viên đầu.
            "hierarchy_candidates": ft.detect_hierarchy(rows, a.first_data_row),
        })
    print(json.dumps(out, ensure_ascii=False, indent=2))


def load_document(path: Path):
    """Đọc functions.json đã có. Trả None nếu chưa có file."""
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as e:
        raise SystemExit(f"{path} không phải JSON hợp lệ ({e}). "
                         "Sửa hoặc xoá file rồi chạy lại.")


def save_document(path: Path, doc: dict) -> None:
    """Ghi nguyên tử: dựng file tạm cùng thư mục rồi replace. Ngắt giữa chừng
    thì bản cũ còn nguyên vẹn, không có file JSON cụt."""
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp = tempfile.mkstemp(dir=str(path.parent), suffix=".tmp")
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            json.dump(doc, f, ensure_ascii=False, indent=2)
            f.write("\n")
        os.replace(tmp, path)
    except BaseException:
        Path(tmp).unlink(missing_ok=True)
        raise


def cmd_write(a) -> None:
    grids = read_grid(a.path, a.sheet)
    name = a.sheet or next(iter(grids))
    grid = grids[name]
    mapping = json.loads(Path(a.mapping).read_text(encoding="utf-8"))

    try:
        tree, skipped = ft.build_tree(grid, mapping)
    except ValueError as e:
        # Lỗi cấu trúc file nguồn: dừng sạch, KHÔNG ghi gì — bản cũ nguyên vẹn.
        raise SystemExit(str(e))
    if not tree:
        raise SystemExit("Không lấy được chức năng nào — kiểm lại ánh xạ cột.")

    out = Path(a.out)
    old_doc = load_document(out)
    old_tree = (old_doc or {}).get("functions") or []
    prev_retired = (old_doc or {}).get("retired_ids") or []

    ft.assign_ids(tree, old_tree, prev_retired)
    ft.carry_status(tree, old_tree)
    retired = ft.compute_retired(old_tree, tree, prev_retired)

    save_document(out, ft.build_document(
        tree, a.system, str(a.path), name, a.date, retired))

    report = {
        "out": str(out),
        "written": sum(1 for n, _ in ft.walk(tree) if not ft.is_use_case(n)),
        "written_use_cases": sum(1 for n, _ in ft.walk(tree) if ft.is_use_case(n)),
        "skipped": skipped,
        "retired": retired,
    }
    if old_doc:
        report["diff"] = ft.diff_trees(old_tree, tree)
    print(json.dumps(report, ensure_ascii=False, indent=2))


def cmd_update(a) -> None:
    """Đổi `status` của một/nhiều FN-ID. Đây là đường DUY NHẤT để code-intel và
    srs-from-code ghi ngược tiến độ — không lệnh nào được sửa tay functions.json.

    Kiểm toàn bộ trước khi ghi: một ID sai là dừng sạch, không ghi phần đúng rồi
    bỏ dở phần sai."""
    path = Path(a.file)
    doc = load_document(path)
    if doc is None:
        raise SystemExit(f"Không thấy {path}. Chạy `fnlist-import` trước.")
    tree = doc.get("functions") or []

    pairs = []
    for item in a.set:
        if "=" not in item:
            raise SystemExit(f"--set '{item}' sai cú pháp, cần dạng FN-01-01=intel.")
        fid, status = item.split("=", 1)
        if status not in ft.STATUSES:
            raise SystemExit(f"Trạng thái '{status}' không hợp lệ. "
                             f"Chỉ nhận: {', '.join(ft.STATUSES)}.")
        node = ft.find_by_id(tree, fid)
        if node is None:
            raise SystemExit(f"Không có {fid} trong {path}.")
        pairs.append((node, fid, status))

    updated = []
    for node, fid, status in pairs:
        old = node.get("status") or "pending"
        if status == "pending":
            node.pop("status", None)
        else:
            node["status"] = status
        updated.append({"id": fid, "cu": old, "moi": status})

    doc["functions"] = [ft.clean_node(n) for n in tree]
    save_document(path, doc)
    print(json.dumps({"file": str(path), "updated": updated},
                     ensure_ascii=False, indent=2))


def main(argv=None):
    p = argparse.ArgumentParser(description="Function list → functions.json")
    sub = p.add_subparsers(dest="cmd", required=True)

    i = sub.add_parser("inspect", help="In cấu trúc thật của file, không đoán cột")
    i.add_argument("path")
    i.add_argument("--sheet", default=None)
    i.add_argument("--max-rows", type=int, default=8, dest="max_rows")
    i.add_argument("--max-cols", type=int, default=12, dest="max_cols")
    i.add_argument("--first-data-row", type=int, default=1, dest="first_data_row",
                   help="Chỉ số 0-based của dòng dữ liệu đầu tiên, dùng khi dò phân cấp")
    i.set_defaults(func=cmd_inspect)

    w = sub.add_parser("write", help="Dựng cây theo ánh xạ cột → functions.json")
    w.add_argument("path")
    w.add_argument("--mapping", required=True, help="File JSON ánh xạ cột")
    w.add_argument("--out", default=".specify/docs/functions.json")
    w.add_argument("--system", default="[TÊN HỆ THỐNG]")
    w.add_argument("--date", required=True, help="Ngày cập nhật YYYY-MM-DD")
    w.add_argument("--sheet", default=None)
    w.set_defaults(func=cmd_write)

    u = sub.add_parser("update", help="Đổi status của FN-ID trong functions.json")
    u.add_argument("--file", default=".specify/docs/functions.json")
    u.add_argument("--set", action="append", required=True, metavar="FN-ID=status",
                   help="vd --set FN-01-01=intel (lặp được nhiều lần)")
    u.set_defaults(func=cmd_update)

    a = p.parse_args(argv)
    a.func(a)


def _needs_openpyxl(argv) -> bool:
    return any(str(x).lower().endswith((".xlsx", ".xlsm")) for x in argv)


def _has_openpyxl() -> bool:
    try:
        import openpyxl  # noqa: F401
        return True
    except ImportError:
        return False


def _bootstrap_and_reexec():
    here = Path(__file__).resolve().parent
    venv = here / ".venv"
    py = venv / ("Scripts" if os.name == "nt" else "bin") / "python"
    if not py.exists():
        print("Lần đầu: đang tạo venv + cài openpyxl...", file=sys.stderr)
        subprocess.check_call([sys.executable, "-m", "venv", str(venv)])
        subprocess.check_call([str(py), "-m", "pip", "install", "-q", "openpyxl"])
    os.execv(str(py), [str(py), str(Path(__file__).resolve()), *sys.argv[1:]])


if __name__ == "__main__":
    if (_needs_openpyxl(sys.argv[1:]) and not _has_openpyxl()
            and os.environ.get("FNLIST_NO_BOOTSTRAP") != "1"):
        _bootstrap_and_reexec()
    main()

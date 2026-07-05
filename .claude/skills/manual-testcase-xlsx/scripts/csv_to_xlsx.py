#!/usr/bin/env python3
"""Chuyển CSV testcase thủ công -> XLSX với format CỐ ĐỊNH.

Dùng: python3 csv_to_xlsx.py <input.csv> [output.xlsx] [--sheet "Tên sheet"]

CSV bắt buộc đúng 14 cột, đúng thứ tự:
ID | Tiêu đề | Nhóm | Ưu tiên | Loại | Tiền điều kiện | Dữ liệu test |
Các bước thực hiện | Kết quả mong đợi | Truy vết |
Kết quả thực tế | Trạng thái | Bug ID | Ghi chú

10 cột đầu = THIẾT KẾ testcase (viết một lần, versioned).
4 cột cuối = THỰC THI (tester điền mỗi test run): thường để trống trong file nguồn.

Format XLSX cố định (không đổi giữa các lần chạy):
- Header nền #1F4E78 chữ trắng đậm, canh giữa, wrap, cao 30.
- Body canh trên + wrap, viền mỏng #BFBFBF.
- Cột Ưu tiên: P1=#C00000, P2=#ED7D31, P3=#A6A6A6 (chữ trắng đậm, canh giữa).
- Cột Trạng thái: dropdown Pass/Fail/Blocked/N/A/Chưa chạy.
- Freeze A2 + auto_filter cả bảng.
- Độ rộng cột: [13,34,20,8,15,34,26,52,58,22, 40,14,14,26].
"""
import sys, csv, os, subprocess


def ensure_openpyxl():
    """Tự bootstrap: nếu thiếu openpyxl thì tạo venv cạnh script, cài openpyxl,
    rồi chạy lại chính script bằng python của venv. Team chỉ cần `python3 csv_to_xlsx.py ...`,
    không phải dựng venv thủ công. Guard bằng biến môi trường để tránh lặp vô hạn."""
    try:
        import openpyxl  # noqa: F401
        return
    except ImportError:
        pass
    if os.environ.get("_TCXLSX_BOOTSTRAPPED"):
        sys.exit("openpyxl vẫn thiếu sau bootstrap. Cài thủ công: pip install openpyxl")
    here = os.path.dirname(os.path.abspath(__file__))
    venv_dir = os.path.join(here, ".venv")
    py = os.path.join(venv_dir, "Scripts" if os.name == "nt" else "bin",
                      "python.exe" if os.name == "nt" else "python")
    if not os.path.exists(py):
        print("Lần đầu chạy: đang tạo venv + cài openpyxl (chỉ 1 lần)...", file=sys.stderr)
        import venv as _venv
        _venv.EnvBuilder(with_pip=True).create(venv_dir)
        subprocess.check_call([py, "-m", "pip", "install", "-q", "openpyxl"])
    env = dict(os.environ, _TCXLSX_BOOTSTRAPPED="1")
    os.execve(py, [py, os.path.abspath(__file__)] + sys.argv[1:], env)


EXPECTED = ["ID","Tiêu đề","Nhóm","Ưu tiên","Loại","Tiền điều kiện",
            "Dữ liệu test","Các bước thực hiện","Kết quả mong đợi","Truy vết",
            "Kết quả thực tế","Trạng thái","Bug ID","Ghi chú"]
WIDTHS = [13,34,20,8,15,34,26,52,58,22, 40,14,14,26]
PRIO = {"P1":"C00000","P2":"ED7D31","P3":"A6A6A6"}
STATUS = ["Pass","Fail","Blocked","N/A","Chưa chạy"]
STATUS_COL = 12  # cột "Trạng thái" (1-based)


INVALID_SHEET = ':\\/?*[]'  # ký tự Excel cấm trong tên sheet


def clean_sheet(name):
    """Bỏ ký tự Excel cấm trong tên sheet (:\\/?*[]) và cắt còn 31 ký tự."""
    for ch in INVALID_SHEET:
        name = name.replace(ch, ' ')
    return ' '.join(name.split())[:31] or "Testcases"


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    sheet = "Testcases"
    if "--sheet" in sys.argv:
        sheet = sys.argv[sys.argv.index("--sheet")+1]
    if not args:
        sys.exit("Thiếu input CSV. Dùng: csv_to_xlsx.py <input.csv> [output.xlsx]")
    src = args[0]
    out = args[1] if len(args) > 1 else os.path.splitext(src)[0] + ".xlsx"

    ensure_openpyxl()  # tự dựng venv + cài openpyxl nếu thiếu, rồi chạy lại
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    rows = list(csv.reader(open(src, encoding="utf-8")))
    if not rows:
        sys.exit("CSV rỗng.")
    if rows[0] != EXPECTED:
        sys.exit("Header CSV sai. Phải đúng 14 cột:\n  " + " | ".join(EXPECTED)
                 + "\nĐang có:\n  " + " | ".join(rows[0]))

    # Chặn dòng sai số field (thừa dấu phẩy -> cột ma / thiếu -> lệch dữ liệu).
    bad = [(i+1, len(r)) for i, r in enumerate(rows) if len(r) != len(EXPECTED)]
    if bad:
        detail = "; ".join(f"dòng {i}={n} field" for i, n in bad)
        sys.exit(f"CSV có {len(bad)} dòng sai số field (phải đúng {len(EXPECTED)}): "
                 f"{detail}. Kiểm tra dấu phẩy thừa/thiếu.")

    wb = Workbook(); ws = wb.active; ws.title = clean_sheet(sheet)
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r in rows:
        ws.append(r)
    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for c in ws[1]:
        c.fill = hdr_fill; c.font = hdr_font; c.border = border
        c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True); c.border = border
        col = PRIO.get(row[3].value)
        if col:
            row[3].fill = PatternFill("solid", fgColor=col)
            row[3].font = Font(color="FFFFFF", bold=True)
            row[3].alignment = Alignment(vertical="top", horizontal="center")

    # Dropdown cho cột Trạng thái (chỉ khi có ít nhất 1 dòng dữ liệu)
    if ws.max_row >= 2:
        dv = DataValidation(type="list", formula1='"%s"' % ",".join(STATUS),
                            allow_blank=True)
        ws.add_data_validation(dv)
        col = get_column_letter(STATUS_COL)
        dv.add(f"{col}2:{col}{ws.max_row}")

    ws.freeze_panes = "A2"
    last = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f"A1:{last}{ws.max_row}"
    wb.save(out)
    print(f"OK: {out} | {ws.max_row-1} cases | {ws.max_column} cols")


if __name__ == "__main__":
    main()
